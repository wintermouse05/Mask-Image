from __future__ import annotations
import io
import os
import tempfile
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .types import ExtractedImage, ImagePlacement


@dataclass
class ImageRef:
    image: XLImage
    anchor_cell: str
    # pixel offsets are not directly available from openpyxl; we work with cell anchors


def _anchor_to_cell(anchor) -> str:
    """Convert an openpyxl anchor to an A1 cell address best-effort."""
    try:
        if hasattr(anchor, "_from") and anchor._from is not None:
            # row/col are 0-based in internal anchor, convert to 1-based
            row = int(getattr(anchor._from, "row", 0)) + 1
            col = int(getattr(anchor._from, "col", 0)) + 1
            return f"{get_column_letter(col)}{row}"
    except Exception:
        pass
    try:
        # Some versions may expose from_ instead
        if hasattr(anchor, "from_") and anchor.from_ is not None:
            row = int(getattr(anchor.from_, "row", 0)) + 1
            col = int(getattr(anchor.from_, "col", 0)) + 1
            return f"{get_column_letter(col)}{row}"
    except Exception:
        pass
    # Fallback string
    try:
        s = str(anchor)
        if "!" in s:
            return s.split("!")[-1]
    except Exception:
        pass
    return "A1"


def _extract_images_from_sheet(ws: Worksheet, tmpdir: str) -> List[ExtractedImage]:
    results: List[ExtractedImage] = []
    idx = 0
    # openpyxl stores images in ws._images
    for img in getattr(ws, '_images', []):
        idx += 1
        anchor = img.anchor
        cell_addr = _anchor_to_cell(anchor)
        # Save image bytes
        data = None
        try:
            # Preferred: XLImage has ref with _data or blob depending on version
            if hasattr(img, "_data") and callable(getattr(img, "_data")):
                data = img._data()
            elif hasattr(img, "ref") and hasattr(img.ref, "_data"):
                data = img.ref._data()  # type: ignore[attr-defined]
            elif hasattr(img, "ref") and hasattr(img.ref, "blob"):
                data = img.ref.blob
            elif hasattr(img, "ref") and hasattr(img.ref, "image"):
                pil = img.ref.image
                buf = io.BytesIO()
                pil.save(buf, format=(getattr(pil, "format", None) or "PNG"))
                data = buf.getvalue()
        except Exception:
            data = None
        if not data:
            # Try saving the embedded image object directly if possible
            try:
                if hasattr(img, "image"):
                    pil = img.image
                    buf = io.BytesIO()
                    pil.save(buf, format=(getattr(pil, "format", None) or "PNG"))
                    data = buf.getvalue()
            except Exception:
                data = None
        if not data:
            # Give up on this image
            continue

        # Infer extension
        ext = 'png'
        if hasattr(img, 'format') and img.format:
            ext = str(img.format).lower()
        elif isinstance(img, XLImage) and getattr(img, 'path', None):
            ext = os.path.splitext(img.path)[1].lstrip('.') or 'png'

        tmp_path = os.path.join(tmpdir, f"{ws.title}_img{idx}.{ext}")
        with open(tmp_path, 'wb') as f:
            f.write(data)

        placement = ImagePlacement(
            sheet_name=ws.title,
            cell=cell_addr,
            left=0,
            top=0,
            width=img.width,
            height=img.height,
            image_id=f"{ws.title}#{idx}",
        )
        results.append(ExtractedImage(placement=placement, image_path=tmp_path))

    return results


def extract_images(input_path: str, sheet_names: Optional[Iterable[str]] = None) -> List[ExtractedImage]:
    wb = load_workbook(input_path)
    tmpdir = tempfile.mkdtemp(prefix='xlimg_')
    results: List[ExtractedImage] = []
    sheets = sheet_names or [ws.title for ws in wb.worksheets]
    for name in sheets:
        ws = wb[name]
        results.extend(_extract_images_from_sheet(ws, tmpdir))
    return results


def write_masked_images(input_path: str, output_path: str, replacements: Dict[str, str]):
    """
    replacements: map from placement.image_id to masked image file path
    """
    wb = load_workbook(input_path)
    for ws in wb.worksheets:
        # Rebuild list, replacing images by position index
        new_images = []
        idx = 0
        for img in getattr(ws, '_images', []):
            idx += 1
            image_id = f"{ws.title}#{idx}"
            if image_id in replacements:
                # Replace with masked image
                masked_path = replacements[image_id]
                new_img = XLImage(masked_path)
                # Preserve size
                new_img.width = img.width
                new_img.height = img.height
                # Preserve anchor
                new_img.anchor = img.anchor
                new_images.append(new_img)
            else:
                new_images.append(img)
        # Assign back (openpyxl doesn't provide a setter; mutate list)
        ws._images = new_images

    # Save to output
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
